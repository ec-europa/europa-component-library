#!/usr/bin/env python3
"""Extract eu-portal-sitemap.xlsx + eu-portal-teasers.docx (in this same
folder) into a single JSON file, for a homepage/inner-page .data.js to
require() as data — see docs/agentic/ecl-static-page.md, Step 2's "Real
content" subsection, and this directory's README.md, for why the derived
JSON isn't hand-retyped and why this script lives next to its source rather
than in a throwaway location.

This script is specific to the two files in this folder (sitemap sheet
shape: "Level 1"/"Level 2" columns; teasers doc shape: Heading-1 category
paragraphs, "List Bullet" items formatted "Title — meta", an optional
following plain paragraph as description, and an optional following
"http..." paragraph as a real destination URL) — it is not a generic
xlsx/docx-to-JSON tool.

Requires openpyxl + python-docx, not repo dependencies — install into a
throwaway venv, e.g.:
    python3 -m venv /tmp/ecl-extract-venv
    /tmp/ecl-extract-venv/bin/pip install openpyxl python-docx
    /tmp/ecl-extract-venv/bin/python extract.py [output_path]

Usage:
    python extract.py [output_path]
        output_path defaults to
        ../../dist/homepage-eu-core/eu-portal-content.json (relative to this
        file) — override it if extracting for a differently-named page.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl
import docx

HERE = Path(__file__).resolve().parent
DEFAULT_OUT = HERE / '..' / '..' / 'dist' / 'homepage-eu-core' / 'eu-portal-content.json'


def extract_sitemap(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Sitemap']
    tree = []
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        l1, l2 = row
        if l1:
            current = {'label': l1, 'children': []}
            tree.append(current)
        if l2 and current is not None:
            current['children'].append(l2)

    notes_ws = wb['Notes']
    notes = {row[0]: row[1] for row in notes_ws.iter_rows(values_only=True) if row[0]}
    return tree, notes


# Bullet pattern: "Title — Type, Date" or "Title — free text"
BULLET_RE = re.compile(r'^(?P<title>.+?)\s+—\s+(?P<rest>.+)$')


def extract_teasers(docx_path):
    d = docx.Document(docx_path)
    teasers = []
    current_category = None
    pending_item = None

    for p in d.paragraphs:
        style = p.style.name if p.style else ''
        text = p.text.strip()
        if not text:
            continue
        if style.startswith('Heading'):
            current_category = text
            continue
        if style == 'List Bullet':
            if pending_item:
                teasers.append(pending_item)
            m = BULLET_RE.match(text)
            if m:
                pending_item = {
                    'category': current_category,
                    'title': m.group('title'),
                    'meta': m.group('rest'),
                    'description': None,
                }
            else:
                pending_item = {'category': current_category, 'title': text, 'meta': None, 'description': None}
            continue
        # Normal paragraph: either a description for the pending bullet, or
        # (Topic/priority, Resource sections) a standalone source URL.
        if text.startswith('http'):
            if pending_item:
                pending_item['url'] = text
        elif not text.startswith('Sources:') and not text.startswith('This is a frozen snapshot'):
            if pending_item and pending_item['description'] is None:
                pending_item['description'] = text
    if pending_item:
        teasers.append(pending_item)

    return teasers


def main():
    out_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_OUT.resolve()

    tree, notes = extract_sitemap(HERE / 'eu-portal-sitemap.xlsx')
    teasers = extract_teasers(HERE / 'eu-portal-teasers.docx')

    data = {'source_notes': notes, 'sitemap': tree, 'teasers': teasers}

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f'Wrote {out_path}')
    print(f'sitemap top-level items: {len(tree)}')
    print(f'teasers extracted: {len(teasers)}')


if __name__ == '__main__':
    main()
